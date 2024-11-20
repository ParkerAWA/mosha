from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook

from zxl import models
from zxl.utils.pagenation import pagination


# Create your views here.


# 部门列表
def department_list(request):
    depart = models.department.objects.all()
    # 分页
    page_object = pagination(request, depart)
    context = {'depart': page_object.page_queryset,
               'page_string': page_object.html()
               }

    return render(request, 'depart/department_list.html', context)


# 部门添加
def depart_add(request):
    if request.method == 'GET':
        return render(request, 'depart/depart_add.html')
    title = request.POST.get('title')
    if not title:
        return render(request, 'depart/depart_add.html', {'error_msg': '部门不能为空!'})
    models.department.objects.create(title=title)
    return redirect('http://127.0.0.1:8000/department/list/')


# 部门删除
def depart_delete(request):
    id = request.GET.get('id')
    models.department.objects.filter(id=id).delete()
    return redirect('http://127.0.0.1:8000/department/list/')


# 部门修改
def depart_update(request, nid):
    if request.method == 'GET':
        row = models.department.objects.filter(id=nid).first()
        # 根据id找的数据
        return render(request, 'depart/depart_update.html', {'row': row})
    title = request.POST.get('title')
    models.department.objects.filter(id=nid).update(title=title)
    # 根据ID对数据进行修改
    return redirect('http://127.0.0.1:8000/department/list/')


"""
批量上传数据(excel)
"""


def depart_upload(request):
    error_message = None
    if request.method == 'GET':
        return render(request, 'depart/depart_upload.html')
    if request.method == 'POST':
        # 1. 获取文件
        file_object = request.FILES.get('exc')
        if not file_object or not file_object.name.endswith(('.xlsx', '.xls')):
            error_message = "格式错误，只支持.xlsx和.xls文件"
        else:
            # 2. 将文件传递到load_workbook()方法中，由load_workbook()读取文件内容
            wb = load_workbook(file_object)
            sheet = wb.worksheets[0]

            # 循环读取数据
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                text = row[0].value
                print(text)
                exists = models.department.objects.filter(title=text).exists()
                if not exists:
                    models.department.objects.create(title=text)

            return redirect('http://127.0.0.1:8000/department/list/')

    # 渲染页面并传递错误信息
    return render(request, 'depart/depart_upload.html', {'error_message': error_message})

# def depart_upload(request):
#     # 1.获取文件
#     file_object = request.FILES.get('exc')
#     if not file_object or not file_object.name.endswith(('.xlsx', '.xls')):
#         return HttpResponse("格式错误，只支持.xlsx和.xls文件")
#     # 2.将文件传递到load_workbook()方法中，由load_workbook()读取文件内容
#     wb = load_workbook(file_object)
#     sheet = wb.worksheets[0]
#     # cell=sheet.cell(1,1)
#
#     # 循环读取数据
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#         text = row[0].value
#         print(text)
#         exists = models.department.objects.filter(title=text).exists()
#         if not exists:
#             models.department.objects.create(title=text)
#
#     return redirect('http://127.0.0.1:8000/department/list/')
